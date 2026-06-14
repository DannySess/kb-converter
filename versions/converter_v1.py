#!/usr/bin/env python3
"""
bms-converter: Converts BMS documents to Markdown + KB settings UI.
"""

import os
import sys
import time
import json
import logging
import threading
import requests
from pathlib import Path
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from flask import Flask, jsonify, render_template_string, request
from flask_socketio import SocketIO, emit

# ── Config ────────────────────────────────────────────────────────────────────
WATCH_DIR   = Path(os.environ.get("WATCH_DIR", "/input"))
OUTPUT_DIR  = Path(os.environ.get("OUTPUT_DIR", "/output"))
CONFIG_PATH = Path(os.environ.get("CONFIG_PATH", "/config/kb_config.json"))
UI_PORT     = int(os.environ.get("UI_PORT", 5000))

CODE_EXTS = {".py", ".js", ".aut", ".tgml", ".xml"}
SKIP_EXTS = {".xbk", ".bak", ".tmp", ".log", ".dat", ".db", ".sqlite", ".idx", ".bin"}
SKIP_DIRS = {"OLD", "old", "Old", "archive", "8_BACKUPS", "Temp", "temp", "V1", "V2", "V3", "V4", "V5", "Update", "WEBPAGES"}

# ── State ─────────────────────────────────────────────────────────────────────
state = {"paused": False, "converted": 0, "skipped": 0, "failed": 0, "current": None, "logs": []}
state_lock = threading.Lock()

app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*")


class SocketHandler(logging.Handler):
    def emit(self, record):
        msg = self.format(record)
        with state_lock:
            state["logs"].append(msg)
            if len(state["logs"]) > 200:
                state["logs"].pop(0)
        socketio.emit("log", {"msg": msg})


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout), SocketHandler()],
)
log = logging.getLogger("bms-converter")

# ── HTML ──────────────────────────────────────────────────────────────────────
HTML = """
<!DOCTYPE html>
<html>
<head>
  <title>BMS Converter</title>
  <script src="https://cdnjs.cloudflare.com/ajax/libs/socket.io/4.7.2/socket.io.min.js"></script>
  <style>
    * { box-sizing: border-box; margin: 0; padding: 0; }
    body { background: #1a1a2e; color: #eee; font-family: monospace; padding: 20px; }
    h1 { color: #00d4ff; margin-bottom: 20px; font-size: 1.4em; }
    .tabs { display:flex; gap:8px; margin-bottom:20px; }
    .tab { padding:8px 20px; border-radius:6px; cursor:pointer; border:none; font-size:0.9em; font-weight:bold; }
    .tab.active { background:#00d4ff; color:#000; }
    .tab:not(.active) { background:#16213e; color:#aaa; }
    .stats { display: flex; gap: 16px; margin-bottom: 20px; flex-wrap: wrap; }
    .stat { background: #16213e; padding: 14px 20px; border-radius: 8px; min-width: 120px; }
    .stat .val { font-size: 2em; font-weight: bold; color: #00d4ff; }
    .stat .label { font-size: 0.75em; color: #aaa; margin-top: 4px; }
    .current { background: #16213e; padding: 12px 16px; border-radius: 8px; margin-bottom: 20px; color: #aaa; font-size: 0.85em; }
    .current span { color: #00d4ff; }
    .controls { margin-bottom: 20px; display:flex; gap:8px; }
    button { padding: 10px 24px; border: none; border-radius: 6px; cursor: pointer; font-size:0.9em; font-weight: bold; }
    #pauseBtn { background: #e94560; color: white; }
    #pauseBtn.paused { background: #00b894; }
    .log-box { background: #0f0f23; border-radius: 8px; padding: 16px; height: 420px; overflow-y: auto; font-size: 0.78em; line-height: 1.6; }
    .log-box .err { color: #e94560; }
    .log-box .ok  { color: #00b894; }
    .log-box .inf { color: #aaa; }
    /* Settings */
    .settings-section { background:#16213e; padding:16px; border-radius:8px; margin-bottom:16px; }
    .settings-section h3 { color:#aaa; font-size:0.85em; margin-bottom:12px; letter-spacing:1px; }
    .input-row { display:flex; gap:8px; margin-bottom:8px; }
    .input-row input { flex:1; background:#0f0f23; border:1px solid #333; color:#eee; padding:8px; border-radius:4px; font-family:monospace; font-size:0.85em; }
    .btn-primary { background:#00d4ff; color:#000; }
    .btn-secondary { background:#333; color:#eee; }
    .btn-success { background:#00b894; color:white; }
    .btn-danger { background:#e94560; color:white; }
    .btn-sm { padding:5px 12px; font-size:0.8em; }
    .mapping-row { display:flex; align-items:center; gap:8px; padding:8px; background:#0f0f23; border-radius:4px; margin-bottom:6px; }
    .folder-name { color:#00d4ff; font-family:monospace; flex:1; }
    .arrow { color:#555; }
    .kb-name { color:#eee; font-family:monospace; flex:1; }
    .kb-id-short { color:#555; font-size:0.7em; flex:1; }
    .unmapped-row { display:flex; align-items:center; gap:8px; padding:8px; background:#0f0f23; border-radius:4px; margin-bottom:6px; }
    .status-msg { font-size:0.8em; margin-top:6px; }
  </style>
</head>
<body>
  <h1>⚙️ BMS Converter</h1>
  
  <div class="tabs">
    <button class="tab active" onclick="showTab('converter')">📊 Converter</button>
    <button class="tab" onclick="showTab('settings')">🗂️ KB Settings</button>
  </div>

  <!-- CONVERTER TAB -->
  <div id="tab-converter">
    <div class="stats">
      <div class="stat"><div class="val" id="converted">0</div><div class="label">Converted</div></div>
      <div class="stat"><div class="val" id="skipped">0</div><div class="label">Skipped</div></div>
      <div class="stat"><div class="val" id="failed" style="color:#e94560">0</div><div class="label">Failed</div></div>
    </div>
    <div class="current">Currently: <span id="current">idle</span></div>
    <div class="controls">
      <button id="pauseBtn" onclick="togglePause()">⏸ Pause</button>
    </div>
    <div class="log-box" id="logBox"></div>
  </div>

  <!-- SETTINGS TAB -->
  <div id="tab-settings" style="display:none">
    <div class="settings-section">
      <h3>OPEN WEBUI CONNECTION</h3>
      <div class="input-row">
        <input id="owui-url" placeholder="http://192.168.0.10:8181">
        <input id="owui-token" type="password" placeholder="sk-...">
        <button class="btn-primary" onclick="saveConnection()">Save</button>
        <button class="btn-secondary" onclick="testConnection()">Test</button>
      </div>
      <div id="conn-status" class="status-msg"></div>
    </div>

    <div class="settings-section">
      <h3>FOLDER → KNOWLEDGE BASE MAPPINGS</h3>
      <div id="mappings-list"><div style="color:#555;font-size:0.85em">Loading...</div></div>
    </div>

    <div class="settings-section">
      <h3>UNMAPPED FOLDERS</h3>
      <div id="unmapped-list"><div style="color:#555;font-size:0.85em">Loading...</div></div>
    </div>
  </div>

  <script>
    const socket = io();
    
    // ── Tab switching ─────────────────────────────────────────
    function showTab(tab) {
      document.getElementById('tab-converter').style.display = tab === 'converter' ? 'block' : 'none';
      document.getElementById('tab-settings').style.display = tab === 'settings' ? 'block' : 'none';
      document.querySelectorAll('.tab').forEach((t, i) => {
        t.classList.toggle('active', (i === 0 && tab === 'converter') || (i === 1 && tab === 'settings'));
      });
      if (tab === 'settings') loadSettings();
    }

    // ── Converter ─────────────────────────────────────────────
    socket.on("log", d => {
      const box = document.getElementById("logBox");
      const line = document.createElement("div");
      const cls = d.msg.includes("ERROR") || d.msg.includes("❌") ? "err"
                : d.msg.includes("✅") ? "ok" : "inf";
      line.className = cls;
      line.textContent = d.msg;
      box.appendChild(line);
      box.scrollTop = box.scrollHeight;
    });

    socket.on("state", d => {
      document.getElementById("converted").textContent = d.converted;
      document.getElementById("skipped").textContent = d.skipped;
      document.getElementById("failed").textContent = d.failed;
      document.getElementById("current").textContent = d.current || "idle";
      const btn = document.getElementById("pauseBtn");
      btn.textContent = d.paused ? "▶ Resume" : "⏸ Pause";
      btn.className = d.paused ? "paused" : "";
    });

    function togglePause() { fetch("/toggle_pause", {method: "POST"}); }

    fetch("/api/state").then(r => r.json()).then(d => {
      document.getElementById("converted").textContent = d.converted;
      document.getElementById("skipped").textContent = d.skipped;
      document.getElementById("failed").textContent = d.failed;
      document.getElementById("current").textContent = d.current || "idle";
      const btn = document.getElementById("pauseBtn");
      btn.textContent = d.paused ? "▶ Resume" : "⏸ Pause";
      btn.className = d.paused ? "paused" : "";
      const box = document.getElementById("logBox");
      d.logs.forEach(msg => {
        const line = document.createElement("div");
        line.className = msg.includes("ERROR") || msg.includes("❌") ? "err" : msg.includes("✅") ? "ok" : "inf";
        line.textContent = msg;
        box.appendChild(line);
      });
      box.scrollTop = box.scrollHeight;
    });

    // ── Settings ──────────────────────────────────────────────
    async function loadSettings() {
      const r = await fetch('/api/settings');
      const d = await r.json();
      document.getElementById('owui-url').value = d.open_webui_url || '';
      document.getElementById('owui-token').value = d.open_webui_token || '';
      renderMappings(d.mappings || {}, d.unmapped || []);
    }

    function renderMappings(mappings, unmapped) {
      const ml = document.getElementById('mappings-list');
      if (Object.keys(mappings).length === 0) {
        ml.innerHTML = '<div style="color:#555;font-size:0.85em">No mappings yet — create KBs from the unmapped folders below.</div>';
      } else {
        ml.innerHTML = Object.entries(mappings).map(([folder, info]) => `
          <div class="mapping-row">
            <span class="folder-name">${folder}/</span>
            <span class="arrow">→</span>
            <span class="kb-name">${info.kb_name}</span>
            <span class="kb-id-short">${info.kb_id.substring(0,8)}...</span>
            <button class="btn-danger btn-sm" onclick="deleteMapping('${folder}')">✕ Remove</button>
          </div>
        `).join('');
      }

      const ul = document.getElementById('unmapped-list');
      if (unmapped.length === 0) {
        ul.innerHTML = '<div style="color:#00b894;font-size:0.85em">✅ All folders mapped!</div>';
      } else {
        ul.innerHTML = unmapped.map(folder => `
          <div class="unmapped-row">
            <span style="color:#e94560;font-family:monospace;flex:1">${folder}/</span>
            <input id="kb-name-${folder}" value="${folder}" 
              style="flex:1;background:#1a1a2e;border:1px solid #333;color:#eee;padding:6px;border-radius:4px;font-family:monospace;font-size:0.85em">
            <button class="btn-success btn-sm" onclick="createAndMap('${folder}')">+ Create KB</button>
          </div>
        `).join('');
      }
    }

    async function saveConnection() {
      const url = document.getElementById('owui-url').value;
      const token = document.getElementById('owui-token').value;
      if (!token || token === '••••••••') {
        // Don't overwrite token if it wasn't changed
        const r = await fetch('/api/settings/connection', {
          method: 'POST', headers: {'Content-Type': 'application/json'},
          body: JSON.stringify({open_webui_url: url})
        });
        const d = await r.json();
        showStatus(d);
      } else {
        const r = await fetch('/api/settings/connection', {
          method: 'POST', headers: {'Content-Type': 'application/json'},
          body: JSON.stringify({open_webui_url: url, open_webui_token: token})
        });
        const d = await r.json();
        showStatus(d);
      }
    }

    async function testConnection() {
      document.getElementById('conn-status').textContent = 'Testing...';
      document.getElementById('conn-status').style.color = '#aaa';
      const r = await fetch('/api/settings/test');
      const d = await r.json();
      showStatus(d);
    }

    function showStatus(d) {
      const el = document.getElementById('conn-status');
      el.textContent = d.message;
      el.style.color = d.ok ? '#00b894' : '#e94560';
    }

    async function createAndMap(folder) {
      const kbName = document.getElementById('kb-name-' + folder).value;
      const btn = document.querySelector(`button[onclick="createAndMap('${folder}')"]`);
      btn.textContent = 'Creating...';
      btn.disabled = true;
      const r = await fetch('/api/settings/create_kb', {
        method: 'POST', headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({folder, kb_name: kbName})
      });
      const d = await r.json();
      if (d.ok) {
        loadSettings();
      } else {
        btn.textContent = '+ Create KB';
        btn.disabled = false;
        alert('Error: ' + d.message);
      }
    }

    async function deleteMapping(folder) {
      if (!confirm('Remove mapping for ' + folder + '?')) return;
      await fetch('/api/settings/mapping/' + encodeURIComponent(folder), {method: 'DELETE'});
      loadSettings();
    }
  </script>
</body>
</html>
"""

# ── Settings helpers ──────────────────────────────────────────────────────────

def load_config():
    if CONFIG_PATH.exists():
        try:
            return json.loads(CONFIG_PATH.read_text())
        except Exception:
            pass
    return {"open_webui_url": "", "open_webui_token": "", "mappings": {}}


def save_config(cfg):
    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    CONFIG_PATH.write_text(json.dumps(cfg, indent=2))


def get_top_folders():
    if not OUTPUT_DIR.exists():
        return []
    return sorted([p.name for p in OUTPUT_DIR.iterdir() if p.is_dir()])


def owui_headers(token):
    return {"Authorization": f"Bearer {token}"}


# ── Flask Routes ──────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/api/state")
def api_state():
    with state_lock:
        return jsonify(state)


@app.route("/toggle_pause", methods=["POST"])
def toggle_pause():
    with state_lock:
        state["paused"] = not state["paused"]
    log.info(f"{'⏸ Converter paused' if state['paused'] else '▶ Converter resumed'}")
    socketio.emit("state", state)
    return jsonify({"paused": state["paused"]})


@app.route("/api/settings")
def api_settings():
    cfg = load_config()
    folders = get_top_folders()
    mapped = set(cfg.get("mappings", {}).keys())
    unmapped = [f for f in folders if f not in mapped]
    return jsonify({
        "open_webui_url": cfg.get("open_webui_url", ""),
        "open_webui_token": cfg.get("open_webui_token", ""),
        "mappings": cfg.get("mappings", {}),
        "unmapped": unmapped
    })


@app.route("/api/settings/connection", methods=["POST"])
def api_save_connection():
    cfg = load_config()
    data = request.json
    cfg["open_webui_url"] = data.get("open_webui_url", cfg.get("open_webui_url", "")).rstrip("/")
    if "open_webui_token" in data and data["open_webui_token"]:
        cfg["open_webui_token"] = data["open_webui_token"]
    save_config(cfg)
    return jsonify({"ok": True, "message": "Connection settings saved ✅"})


@app.route("/api/settings/test")
def api_test_connection():
    cfg = load_config()
    url = cfg.get("open_webui_url", "")
    token = cfg.get("open_webui_token", "")
    if not url or not token:
        return jsonify({"ok": False, "message": "URL and token not configured"})
    try:
        r = requests.get(f"{url}/api/v1/auths/", headers=owui_headers(token), timeout=5)
        if r.status_code == 200:
            return jsonify({"ok": True, "message": f"✅ Connected to Open WebUI successfully"})
        else:
            return jsonify({"ok": False, "message": f"❌ HTTP {r.status_code} — check your token"})
    except Exception as e:
        return jsonify({"ok": False, "message": f"❌ Cannot reach Open WebUI: {e}"})


@app.route("/api/settings/create_kb", methods=["POST"])
def api_create_kb():
    cfg = load_config()
    url = cfg.get("open_webui_url", "")
    token = cfg.get("open_webui_token", "")
    if not url or not token:
        return jsonify({"ok": False, "message": "Configure Open WebUI connection first"})
    
    data = request.json
    folder = data.get("folder", "")
    kb_name = data.get("kb_name", folder)
    
    try:
        r = requests.post(
            f"{url}/api/v1/knowledge/create",
            headers={**owui_headers(token), "Content-Type": "application/json"},
            json={"name": kb_name, "description": f"BMS documents from {folder}/"},
            timeout=10
        )
        r.raise_for_status()
        kb_id = r.json()["id"]
        
        if "mappings" not in cfg:
            cfg["mappings"] = {}
        cfg["mappings"][folder] = {"kb_id": kb_id, "kb_name": kb_name}
        save_config(cfg)
        
        log.info(f"✅ Created KB '{kb_name}' ({kb_id}) for folder {folder}/")
        return jsonify({"ok": True, "kb_id": kb_id, "message": f"KB '{kb_name}' created and mapped!"})
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)})


@app.route("/api/settings/mapping/<folder>", methods=["DELETE"])
def api_delete_mapping(folder):
    cfg = load_config()
    if folder in cfg.get("mappings", {}):
        del cfg["mappings"][folder]
        save_config(cfg)
    return jsonify({"ok": True})


# ── Conversion ────────────────────────────────────────────────────────────────

def broadcast_state():
    with state_lock:
        socketio.emit("state", dict(state))


def output_path(input_path: Path) -> Path:
    rel = input_path.relative_to(WATCH_DIR)
    if rel.suffix.lower() == ".md":
        return OUTPUT_DIR / rel
    return OUTPUT_DIR / rel.parent / (rel.name + ".md")


def already_converted(input_path: Path) -> bool:
    out = output_path(input_path)
    if not out.exists():
        return False
    return out.stat().st_mtime >= input_path.stat().st_mtime


def fix_spacing(text: str) -> str:
    import re
    text = re.sub(r'([a-z])([A-Z])', r'\1 \2', text)
    text = re.sub(r'([a-zA-Z])(\d)', r'\1 \2', text)
    text = re.sub(r'(\d)([a-zA-Z])', r'\1 \2', text)
    text = re.sub(r'([.,;:])([A-Za-z])', r'\1 \2', text)
    return text


def convert_pdf(path: Path) -> str:
    try:
        import pdfplumber
        md = f"# {path.stem}\n\n"
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        if not table:
                            continue
                        header = [str(c) if c is not None else "" for c in table[0]]
                        md += "| " + " | ".join(header) + " |\n"
                        md += "| " + " | ".join(["---"] * len(header)) + " |\n"
                        for row in table[1:]:
                            cells = [str(c) if c is not None else "" for c in row]
                            md += "| " + " | ".join(cells) + " |\n"
                        md += "\n"
                else:
                    text = page.extract_text()
                    if text:
                        md += fix_spacing(text) + "\n\n"
        return md
    except Exception as e:
        log.error(f"PDF conversion failed for {path.name}: {e}")
        return f"# {path.stem}\n\n*Conversion failed: {e}*\n"


def convert_xlsx(path: Path) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        md = f"# {path.stem}\n\n"
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            md += f"## {sheet_name}\n\n"
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c) if c is not None else "" for c in rows[0]]
            md += "| " + " | ".join(header) + " |\n"
            md += "| " + " | ".join(["---"] * len(header)) + " |\n"
            for row in rows[1:]:
                cells = [str(c) if c is not None else "" for c in row]
                md += "| " + " | ".join(cells) + " |\n"
            md += "\n"
        return md
    except Exception as e:
        log.error(f"XLSX conversion failed for {path.name}: {e}")
        return f"# {path.stem}\n\n*Conversion failed: {e}*\n"


def convert_xls(path: Path) -> str:
    try:
        import xlrd
        wb = xlrd.open_workbook(path)
        md = f"# {path.stem}\n\n"
        for sheet in wb.sheets():
            md += f"## {sheet.name}\n\n"
            if sheet.nrows == 0:
                continue
            header = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
            md += "| " + " | ".join(header) + " |\n"
            md += "| " + " | ".join(["---"] * len(header)) + " |\n"
            for r in range(1, sheet.nrows):
                row = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
                md += "| " + " | ".join(row) + " |\n"
            md += "\n"
        return md
    except Exception as e:
        log.error(f"XLS conversion failed for {path.name}: {e}")
        return f"# {path.stem}\n\n*Conversion failed: {e}*\n"


def convert_csv(path: Path) -> str:
    try:
        import pandas as pd
        df = pd.read_csv(path, dtype=str).fillna("")
        md = f"# {path.stem}\n\n"
        md += df.to_markdown(index=False)
        md += "\n"
        return md
    except Exception as e:
        log.error(f"CSV conversion failed for {path.name}: {e}")
        return f"# {path.stem}\n\n*Conversion failed: {e}*\n"


def convert_code(path: Path) -> str:
    ext_map = {".py": "python", ".js": "javascript", ".aut": "text", ".tgml": "xml", ".xml": "xml"}
    lang = ext_map.get(path.suffix.lower(), "text")
    try:
        content = path.read_text(errors="replace")
        return f"# {path.stem}\n\n```{lang}\n{content}\n```\n"
    except Exception as e:
        return f"# {path.stem}\n\n*Read failed: {e}*\n"


def convert_file(path: Path) -> bool:
    with state_lock:
        if state["paused"]:
            return False

    ext = path.suffix.lower()
    if ext in SKIP_EXTS:
        return False
    for skip in SKIP_DIRS:
        if skip in path.parts:
            return False
    if already_converted(path):
        with state_lock:
            state["skipped"] += 1
        broadcast_state()
        return False

    with state_lock:
        state["current"] = path.name
    broadcast_state()

    log.info(f"🔄 Converting: {path.relative_to(WATCH_DIR)}")

    if ext == ".pdf":
        content = convert_pdf(path)
    elif ext in {".xlsx", ".xlsm"}:
        content = convert_xlsx(path)
    elif ext == ".xls":
        content = convert_xls(path)
    elif ext == ".csv":
        content = convert_csv(path)
    elif ext in CODE_EXTS:
        content = convert_code(path)
    elif ext == ".md":
        content = path.read_text(errors="replace")
    else:
        return False

    out = output_path(path)
    out.parent.mkdir(parents=True, exist_ok=True)

    try:
        out.write_text(content, encoding="utf-8")
        log.info(f"  ✅ Saved: {out.relative_to(OUTPUT_DIR)}")
        with state_lock:
            state["converted"] += 1
            state["current"] = None
        broadcast_state()
        return True
    except Exception as e:
        log.error(f"  ❌ Write failed: {e}")
        with state_lock:
            state["failed"] += 1
            state["current"] = None
        broadcast_state()
        return False


class ConvertHandler(FileSystemEventHandler):
    def on_created(self, event):
        if event.is_directory: return
        time.sleep(2)
        convert_file(Path(event.src_path))

    def on_modified(self, event):
        if event.is_directory: return
        convert_file(Path(event.src_path))

    def on_moved(self, event):
        if event.is_directory: return
        convert_file(Path(event.dest_path))


def startup_scan():
    all_files = [p for p in WATCH_DIR.rglob("*") if p.is_file()]
    log.info(f"🔍 Startup scan: {len(all_files)} file(s)...")
    for path in sorted(all_files):
        convert_file(path)
    log.info("✅ Startup scan complete.")


def converter_thread():
    if not WATCH_DIR.exists():
        log.error(f"Input directory does not exist: {WATCH_DIR}")
        return
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    log.info("=" * 60)
    log.info("bms-converter starting")
    log.info(f"  Input:   {WATCH_DIR}")
    log.info(f"  Output:  {OUTPUT_DIR}")
    log.info(f"  Config:  {CONFIG_PATH}")
    log.info(f"  UI:      http://0.0.0.0:{UI_PORT}")
    log.info("=" * 60)
    startup_scan()
    handler = ConvertHandler()
    observer = Observer()
    observer.schedule(handler, str(WATCH_DIR), recursive=True)
    observer.start()
    log.info(f"👁  Watching {WATCH_DIR} ...")
    try:
        while True:
            time.sleep(5)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()


if __name__ == "__main__":
    t = threading.Thread(target=converter_thread, daemon=True)
    t.start()
    socketio.run(app, host="0.0.0.0", port=UI_PORT, allow_unsafe_werkzeug=True)
