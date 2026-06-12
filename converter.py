#!/usr/bin/env python3
"""
bms-converter: Converts BMS documents to Markdown with a web UI.
"""

import os
import sys
import time
import logging
import threading
from pathlib import Path
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from flask import Flask, jsonify, render_template_string
from flask_socketio import SocketIO, emit

# ── Config ────────────────────────────────────────────────────────────────────
WATCH_DIR  = Path(os.environ.get("WATCH_DIR", "/input"))
OUTPUT_DIR = Path(os.environ.get("OUTPUT_DIR", "/output"))
UI_PORT    = int(os.environ.get("UI_PORT", 5000))

CODE_EXTS = {".py", ".js", ".aut", ".tgml", ".xml"}
SKIP_EXTS = {".xbk", ".bak", ".tmp", ".log", ".dat", ".db", ".sqlite", ".idx", ".bin"}
SKIP_DIRS = {"OLD", "old", "archive", "8_BACKUPS", "Temp", "temp"}

# ── State ─────────────────────────────────────────────────────────────────────
state = {
    "paused": False,
    "converted": 0,
    "skipped": 0,
    "failed": 0,
    "current": None,
    "logs": [],
}
state_lock = threading.Lock()

# ── Logging ───────────────────────────────────────────────────────────────────
app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*")

class SocketHandler(logging.Handler):
    def emit(self, record):
        msg = self.format(record)
        with state_lock:
            state["logs"].append(msg)
            if len(state["logs"]) > 200:
                state["logs"].pop(0)
        socketio.emit("log", {"msg": msg})

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout), SocketHandler()],
)
log = logging.getLogger("bms-converter")

# ── Web UI ────────────────────────────────────────────────────────────────────
HTML = """
<!DOCTYPE html>
<html>
<head>
  <title>BMS Converter</title>
  <script src="https://cdnjs.cloudflare.com/ajax/libs/socket.io/4.7.2/socket.io.min.js"></script>
  <style>
    * { box-sizing: border-box; margin: 0; padding: 0; }
    body { background: #1a1a2e; color: #eee; font-family: monospace; padding: 20px; }
    h1 { color: #00d4ff; margin-bottom: 20px; font-size: 1.4em; }
    .stats { display: flex; gap: 16px; margin-bottom: 20px; flex-wrap: wrap; }
    .stat { background: #16213e; padding: 14px 20px; border-radius: 8px; min-width: 120px; }
    .stat .val { font-size: 2em; font-weight: bold; color: #00d4ff; }
    .stat .label { font-size: 0.75em; color: #aaa; margin-top: 4px; }
    .current { background: #16213e; padding: 12px 16px; border-radius: 8px; margin-bottom: 20px; color: #aaa; font-size: 0.85em; }
    .current span { color: #00d4ff; }
    .controls { margin-bottom: 20px; }
    button { padding: 10px 24px; border: none; border-radius: 6px; cursor: pointer; font-size: 1em; font-weight: bold; }
    #pauseBtn { background: #e94560; color: white; }
    #pauseBtn.paused { background: #00b894; }
    .log-box { background: #0f0f23; border-radius: 8px; padding: 16px; height: 420px; overflow-y: auto; font-size: 0.78em; line-height: 1.6; }
    .log-box .err { color: #e94560; }
    .log-box .ok  { color: #00b894; }
    .log-box .inf { color: #aaa; }
  </style>
</head>
<body>
  <h1>⚙️ BMS Converter</h1>
  <div class="stats">
    <div class="stat"><div class="val" id="converted">0</div><div class="label">Converted</div></div>
    <div class="stat"><div class="val" id="skipped">0</div><div class="label">Skipped</div></div>
    <div class="stat"><div class="val" id="failed" style="color:#e94560">0</div><div class="label">Failed</div></div>
  </div>
  <div class="current">Currently: <span id="current">idle</span></div>
  <div class="controls">
    <button id="pauseBtn" onclick="togglePause()">⏸ Pause</button>
  </div>
  <div class="log-box" id="logBox"></div>

  <script>
    const socket = io();
    let paused = false;

    socket.on("log", d => {
      const box = document.getElementById("logBox");
      const line = document.createElement("div");
      const cls = d.msg.includes("ERROR") || d.msg.includes("❌") ? "err"
                : d.msg.includes("✅") ? "ok" : "inf";
      line.className = cls;
      line.textContent = d.msg;
      box.appendChild(line);
      box.scrollTop = box.scrollHeight;
    });

    socket.on("state", d => {
      document.getElementById("converted").textContent = d.converted;
      document.getElementById("skipped").textContent = d.skipped;
      document.getElementById("failed").textContent = d.failed;
      document.getElementById("current").textContent = d.current || "idle";
      paused = d.paused;
      const btn = document.getElementById("pauseBtn");
      btn.textContent = paused ? "▶ Resume" : "⏸ Pause";
      btn.className = paused ? "paused" : "";
    });

    function togglePause() {
      fetch("/toggle_pause", {method: "POST"});
    }

    // Load initial state and logs
    fetch("/api/state").then(r => r.json()).then(d => {
      document.getElementById("converted").textContent = d.converted;
      document.getElementById("skipped").textContent = d.skipped;
      document.getElementById("failed").textContent = d.failed;
      document.getElementById("current").textContent = d.current || "idle";
      paused = d.paused;
      const btn = document.getElementById("pauseBtn");
      btn.textContent = paused ? "▶ Resume" : "⏸ Pause";
      btn.className = paused ? "paused" : "";
      const box = document.getElementById("logBox");
      d.logs.forEach(msg => {
        const line = document.createElement("div");
        const cls = msg.includes("ERROR") || msg.includes("❌") ? "err"
                  : msg.includes("✅") ? "ok" : "inf";
        line.className = cls;
        line.textContent = msg;
        box.appendChild(line);
      });
      box.scrollTop = box.scrollHeight;
    });
  </script>
</body>
</html>
"""

@app.route("/")
def index():
    return render_template_string(HTML)

@app.route("/api/state")
def api_state():
    with state_lock:
        return jsonify(state)

@app.route("/toggle_pause", methods=["POST"])
def toggle_pause():
    with state_lock:
        state["paused"] = not state["paused"]
        status = "paused" if state["paused"] else "resumed"
    log.info(f"{'⏸ Converter paused' if state['paused'] else '▶ Converter resumed'}")
    socketio.emit("state", state)
    return jsonify({"paused": state["paused"]})

def broadcast_state():
    with state_lock:
        socketio.emit("state", dict(state))

# ── Conversion ────────────────────────────────────────────────────────────────
def output_path(input_path: Path) -> Path:
    rel = input_path.relative_to(WATCH_DIR)
    return OUTPUT_DIR / rel.with_suffix(".md")

def already_converted(input_path: Path) -> bool:
    out = output_path(input_path)
    if not out.exists():
        return False
    return out.stat().st_mtime >= input_path.stat().st_mtime

def fix_spacing(text: str) -> str:
    import re
    text = re.sub(r"([a-z])([A-Z])", r"\1 \2", text)
    text = re.sub(r"([a-zA-Z])(\d)", r"\1 \2", text)
    text = re.sub(r"(\d)([a-zA-Z])", r"\1 \2", text)
    text = re.sub(r"([.,;:])([A-Za-z])", r"\1 \2", text)
    return text

def convert_pdf(path: Path) -> str:
    try:
        import pdfplumber
        md = f"# {path.stem}\n\n"
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                # Extract tables first
                tables = page.extract_tables()
                table_bboxes = [t.bbox for t in page.find_tables()] if tables else []
                if tables:
                    for table in tables:
                        if not table:
                            continue
                        header = [str(c) if c is not None else "" for c in table[0]]
                        md += "| " + " | ".join(header) + " |\n"
                        md += "| " + " | ".join(["---"] * len(header)) + " |\n"
                        for row in table[1:]:
                            cells = [str(c) if c is not None else "" for c in row]
                            md += "| " + " | ".join(cells) + " |\n"
                        md += "\n"
                # Extract text outside tables
                text = page.extract_text()
                if text:
                    md += fix_spacing(text) + "\n\n"
        return md
    except Exception as e:
        log.error(f"PDF conversion failed for {path.name}: {e}")
        return f"# {path.stem}\n\n*Conversion failed: {e}*\n"

def convert_xlsx(path: Path) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        md = f"# {path.stem}\n\n"
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            md += f"## {sheet_name}\n\n"
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c) if c is not None else "" for c in rows[0]]
            md += "| " + " | ".join(header) + " |\n"
            md += "| " + " | ".join(["---"] * len(header)) + " |\n"
            for row in rows[1:]:
                cells = [str(c) if c is not None else "" for c in row]
                md += "| " + " | ".join(cells) + " |\n"
            md += "\n"
        return md
    except Exception as e:
        log.error(f"XLSX conversion failed for {path.name}: {e}")
        return f"# {path.stem}\n\n*Conversion failed: {e}*\n"

def convert_xls(path: Path) -> str:
    try:
        import xlrd
        wb = xlrd.open_workbook(path)
        md = f"# {path.stem}\n\n"
        for sheet in wb.sheets():
            md += f"## {sheet.name}\n\n"
            if sheet.nrows == 0:
                continue
            header = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
            md += "| " + " | ".join(header) + " |\n"
            md += "| " + " | ".join(["---"] * len(header)) + " |\n"
            for r in range(1, sheet.nrows):
                row = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
                md += "| " + " | ".join(row) + " |\n"
            md += "\n"
        return md
    except Exception as e:
        log.error(f"XLS conversion failed for {path.name}: {e}")
        return f"# {path.stem}\n\n*Conversion failed: {e}*\n"

def convert_csv(path: Path) -> str:
    try:
        import pandas as pd
        df = pd.read_csv(path, dtype=str).fillna("")
        md = f"# {path.stem}\n\n"
        md += df.to_markdown(index=False)
        md += "\n"
        return md
    except Exception as e:
        log.error(f"CSV conversion failed for {path.name}: {e}")
        return f"# {path.stem}\n\n*Conversion failed: {e}*\n"

def convert_code(path: Path) -> str:
    ext_map = {".py": "python", ".js": "javascript", ".aut": "text", ".tgml": "xml", ".xml": "xml"}
    lang = ext_map.get(path.suffix.lower(), "text")
    try:
        content = path.read_text(errors="replace")
        return f"# {path.stem}\n\n```{lang}\n{content}\n```\n"
    except Exception as e:
        return f"# {path.stem}\n\n*Read failed: {e}*\n"

def convert_file(path: Path) -> bool:
    with state_lock:
        if state["paused"]:
            return False

    ext = path.suffix.lower()
    if ext in SKIP_EXTS:
        return False
    for skip in SKIP_DIRS:
        if skip in path.parts:
            return False
    if already_converted(path):
        with state_lock:
            state["skipped"] += 1
        broadcast_state()
        return False

    with state_lock:
        state["current"] = path.name
    broadcast_state()

    log.info(f"🔄 Converting: {path.relative_to(WATCH_DIR)}")

    if ext == ".pdf":
        content = convert_pdf(path)
    elif ext in {".xlsx", ".xlsm"}:
        content = convert_xlsx(path)
    elif ext == ".xls":
        content = convert_xls(path)
    elif ext == ".csv":
        content = convert_csv(path)
    elif ext in CODE_EXTS:
        content = convert_code(path)
    elif ext == ".md":
        content = path.read_text(errors="replace")
    else:
        return False

    out = output_path(path)
    out.parent.mkdir(parents=True, exist_ok=True)

    try:
        out.write_text(content, encoding="utf-8")
        log.info(f"  ✅ Saved: {out.relative_to(OUTPUT_DIR)}")
        with state_lock:
            state["converted"] += 1
            state["current"] = None
        broadcast_state()
        return True
    except Exception as e:
        log.error(f"  ❌ Write failed: {e}")
        with state_lock:
            state["failed"] += 1
            state["current"] = None
        broadcast_state()
        return False

class ConvertHandler(FileSystemEventHandler):
    def on_created(self, event):
        if event.is_directory: return
        time.sleep(2)
        convert_file(Path(event.src_path))
    def on_modified(self, event):
        if event.is_directory: return
        convert_file(Path(event.src_path))
    def on_moved(self, event):
        if event.is_directory: return
        convert_file(Path(event.dest_path))

def startup_scan():
    all_files = [p for p in WATCH_DIR.rglob("*") if p.is_file()]
    log.info(f"🔍 Startup scan: found {len(all_files)} file(s)...")
    for path in sorted(all_files):
        convert_file(path)
    log.info(f"✅ Startup scan complete.")

def converter_thread():
    if not WATCH_DIR.exists():
        log.error(f"Input directory does not exist: {WATCH_DIR}")
        return
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    log.info("=" * 60)
    log.info("bms-converter starting")
    log.info(f"  Input:   {WATCH_DIR}")
    log.info(f"  Output:  {OUTPUT_DIR}")
    log.info(f"  UI:      http://0.0.0.0:{UI_PORT}")
    log.info("=" * 60)
    startup_scan()
    handler = ConvertHandler()
    observer = Observer()
    observer.schedule(handler, str(WATCH_DIR), recursive=True)
    observer.start()
    log.info(f"👁  Watching {WATCH_DIR} for changes...")
    try:
        while True:
            time.sleep(5)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

if __name__ == "__main__":
    t = threading.Thread(target=converter_thread, daemon=True)
    t.start()
    socketio.run(app, host="0.0.0.0", port=UI_PORT, allow_unsafe_werkzeug=True)
